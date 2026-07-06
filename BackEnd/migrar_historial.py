"""
Script de migración: Historial_xlsx -> Data_FTP JSON
Leer y ejecutar una sola vez desde la carpeta BackEnd.

  python migrar_historial.py
"""
import sys, json
from pathlib import Path
import openpyxl

# ── Rutas base ────────────────────────────────────────────────────────────────
BASE_DIR          = Path(__file__).resolve().parent
DATA_DIR          = BASE_DIR.parent / "Data"
RUTA_HISTORIAL    = DATA_DIR / "FTP" / "Historial_xlsx"
RUTA_DATA_FTP     = DATA_DIR / "FTP" / "Data_FTP"
RUTA_INDICADORES  = DATA_DIR / "FTP" / "Mapeo_json"

MESES_NOMBRES = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
]
MESES_LISTA = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

# Indicadores que pertenecen a otro sistema (IN_ASS, Px_Referidos, etc.)
EXCLUIR_CATEGORIAS = {"IN_ASS", "Px_Referidos"}


def cargar_semaforo(indicador: str) -> dict:
    """Lee el semáforo del JSON de configuración del indicador."""
    categoria = indicador.split()[0]
    ruta_json = RUTA_INDICADORES / f"{categoria}.json"
    if not ruta_json.exists():
        return {}
    try:
        with open(ruta_json, encoding="utf-8") as f:
            data = json.load(f)
        # Buscar el indicador dentro del JSON de categoría
        for ind in data.get("indicadores", {}).values():
            if ind.get("id") == indicador or ind.get("nombre") == indicador:
                return ind.get("semaforo", {})
        # Si el JSON tiene el indicador directamente
        return data.get("semaforo", {})
    except Exception:
        return {}


def calcular_color(valor, idx_mes: int, semaforo: dict) -> str:
    """Calcula el color del semáforo para un valor dado."""
    if valor is None or valor == "":
        return "Gris"
    try:
        val     = float(valor)
        limites = semaforo.get(MESES_LISTA[idx_mes], semaforo)
        if not limites:
            return "Gris"
        v_esp      = limites.get("Esperado", 0)
        tiene_alto = "Alto" in limites
        v_critico  = limites.get("Alto") if tiene_alto else limites.get("Bajo", 0)

        if tiene_alto:
            if val <= v_esp:      return "Verde"
            elif val < v_critico: return "Amarillo"
            else:                 return "Rojo"
        else:
            if val >= v_esp:      return "Verde"
            elif val > v_critico: return "Amarillo"
            else:                 return "Rojo"
    except Exception:
        return "Gris"


def migrar_excel(ruta_excel: Path, indicador: str, ano: str, semaforo: dict) -> int:
    """
    Lee un Excel histórico y guarda sus datos en el JSON correspondiente.
    Retorna el número de meses migrados.
    """
    try:
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"  ERROR No se pudo abrir {ruta_excel.name}: {e}")
        return 0

    # Leer unidades desde columna 1 (filas 11+)
    unidades = []
    fila = 11
    while True:
        val = ws.cell(row=fila, column=1).value
        if val is None:
            break
        unidades.append(str(val).strip())
        fila += 1

    if not unidades:
        wb.close()
        print(f"  ERROR Sin unidades detectadas en {ruta_excel.name}")
        return 0

    # Destino JSON
    categoria = indicador.split()[0]
    nombre_json = indicador.replace(" ", "_") + ".json"
    ruta_json   = RUTA_DATA_FTP / str(ano) / categoria / nombre_json

    # Leer JSON existente o crear nuevo
    if ruta_json.exists():
        with open(ruta_json, encoding="utf-8") as f:
            contenido = json.load(f)
    else:
        contenido = {"ANIO": int(ano), "MESES": {}}

    meses_migrados = 0

    for idx_mes in range(12):
        col_num = (idx_mes * 3) + 2
        datos_mes = {}
        tiene_dato = False

        for idx_u, unidad in enumerate(unidades):
            num = ws.cell(row=11 + idx_u, column=col_num).value
            den = ws.cell(row=11 + idx_u, column=col_num + 1).value
            pct = ws.cell(row=11 + idx_u, column=col_num + 2).value

            if num is None and den is None and pct is None:
                continue
            tiene_dato = True
            color = calcular_color(pct, idx_mes, semaforo)
            datos_mes[unidad] = {
                "numerador":   num,
                "denominador": den,
                "%":           pct,
                "color":       color,
            }

        if tiene_dato:
            mes_nombre = MESES_NOMBRES[idx_mes]
            # El Excel es la fuente de verdad — siempre sobreescribir
            contenido["MESES"][mes_nombre] = datos_mes
            meses_migrados += 1

    wb.close()

    if meses_migrados > 0:
        ruta_json.parent.mkdir(parents=True, exist_ok=True)
        with open(ruta_json, "w", encoding="utf-8") as f:
            json.dump(contenido, f, ensure_ascii=False, indent=4)

    return meses_migrados


def main():
    print("=" * 60)
    print("Migracion Historial_xlsx -> Data_FTP JSON")
    print("=" * 60)

    total_archivos = 0
    total_meses    = 0
    errores        = []

    # Recorrer carpetas: {categoria}/{indicador}/{ano}/
    for cat_dir in sorted(RUTA_HISTORIAL.iterdir()):
        if not cat_dir.is_dir():
            continue
        categoria = cat_dir.name
        if categoria in EXCLUIR_CATEGORIAS:
            print(f"\n[OMITIDO] {categoria}")
            continue

        for ind_dir in sorted(cat_dir.iterdir()):
            if not ind_dir.is_dir():
                continue
            indicador = ind_dir.name

            for ano_dir in sorted(ind_dir.iterdir()):
                if not ano_dir.is_dir():
                    continue
                ano = ano_dir.name
                if not ano.isdigit():
                    continue

                # Buscar el Excel principal (que contiene el año en el nombre)
                excels = [f for f in ano_dir.glob("*.xlsx") if ano in f.stem]
                if not excels:
                    continue

                ruta_excel = excels[0]
                semaforo   = cargar_semaforo(indicador)
                print(f"\n{indicador} [{ano}]  ->  {ruta_excel.name}")

                n = migrar_excel(ruta_excel, indicador, ano, semaforo)
                if n > 0:
                    print(f"  OK {n} mes(es) migrados")
                    total_meses    += n
                    total_archivos += 1
                else:
                    print(f"  - Sin datos nuevos")

    print("\n" + "=" * 60)
    print(f"Archivos procesados : {total_archivos}")
    print(f"Meses migrados      : {total_meses}")
    print("=" * 60)


if __name__ == "__main__":
    main()

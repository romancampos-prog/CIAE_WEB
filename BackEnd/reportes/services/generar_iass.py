"""
Módulo  : generar_IASS.py
Carpeta : reportes/services/
Qué hace: Genera el Excel completo IASS (01-06) con históricos, semáforo y acumulados.
Usado en: indicadores_IASS/services/procesar_service.py, reportes/controllers/reportes_controller.py
"""
import json
import io
import xlsxwriter
import openpyxl  # sigue usándose para guardar el xlsx resultante
from configs.unidades import MESES, ORDEN_IASS01, ORDEN_DEMAS_IASS, UNIDADES_HGS_IASS01
from configs.settings import RUTA_IASS_JSON, RUTA_DATA_IASS

with open(RUTA_IASS_JSON, encoding="utf-8") as _f:
    _CFG = json.load(_f)

RUTA_BASE_IASS = RUTA_DATA_IASS

UNIDADES_IASS     = ORDEN_IASS01 + ["DELEGACION"]
UNIDADES_UCI      = ORDEN_DEMAS_IASS
_UNIDADES_HGS_SET = set(UNIDADES_HGS_IASS01)


def _color_tasa_01(tasa, unidad):
    sem      = _CFG["IASS 01"]["Semaforo"]
    umbrales = sem.get("HGS") if unidad in _UNIDADES_HGS_SET else sem.get("Otros")
    if not umbrales or tasa is None:
        return "Rojo"
    esp = umbrales.get("Esperado", {})
    med = umbrales.get("Medio", {})
    if esp.get("Mayor", 0) <= tasa <= esp.get("Menor", 0):
        return "Verde"
    elif med.get("Mayor", 0) <= tasa < med.get("Menor", 0):
        return "Amarillo"
    return "Rojo"


def _color_tasa_uci(tasa, indicador):
    sem = _CFG[indicador].get("Semaforo", {})
    if tasa is None:
        return "Rojo"
    esp = sem.get("Esperado", {})
    med = sem.get("Medio", {})
    if tasa > esp.get("Menor", 0):
        return "Rojo"
    elif tasa >= esp.get("Mayor", 0):
        return "Verde"
    elif tasa >= med.get("Mayor", 0):
        return "Amarillo"
    return "Rojo"


_NOMBRE_A_NUM = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}


def _leer_historicos_IASS(anio: str, mes_num: int) -> dict:
    """
    Lee los datos históricos de todos los indicadores desde los JSON por indicador,
    excluyendo el mes actual (mes_num). Reemplaza la lectura desde Excel.
    Devuelve: {ind_key: {mes_num_int: {unit: {numerador, denominador, tasa, color}}}}
    """
    ruta_sesion = RUTA_DATA_IASS / str(anio)
    historicos  = {}

    for ind_n in range(1, 7):
        ind_key   = f"IASS 0{ind_n}"
        ruta_json = ruta_sesion / f"IASS_0{ind_n}.json"
        if not ruta_json.exists():
            historicos[ind_key] = {}
            continue

        try:
            with open(ruta_json, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            print(f"[IASS JSON] Error leyendo {ruta_json}: {e}")
            historicos[ind_key] = {}
            continue

        h_ind = {}
        for mes_nombre, mes_data in data.get("MESES", {}).items():
            m = _NOMBRE_A_NUM.get(mes_nombre.upper())
            if m is None or m == mes_num:
                continue
            datos_mes = {
                unit: {
                    "numerador":   v.get("NUMERADOR"),
                    "denominador": v.get("DENOMINADOR"),
                    "tasa":        v.get("TASA"),
                    "color":       (v.get("COLOR") or "Rojo").capitalize(),
                }
                for unit, v in mes_data.get("DATOS", {}).items()
            }
            if datos_mes:
                h_ind[m] = datos_mes
        historicos[ind_key] = h_ind

    total = sum(len(v) for v in historicos.values())
    print(f"[IASS JSON] {total} meses de datos cargados")
    return historicos


def _escribir_unidades(hoja, estilos, fila_inicio, columna):
    for unidad in UNIDADES_IASS:
        estilo = estilos["delegacion_txt"] if unidad == "DELEGACIÓN" else estilos["lista_unidades_txt"]
        hoja.write(fila_inicio, columna, unidad, estilo)
        fila_inicio += 1


def _escribir_bloque_mesIASS_01(hoja, estilos, fila_titulo_mes, fila_subencabezado, col_inicio, cantidad_unidades, nombre_mes):
    hoja.merge_range(fila_titulo_mes, col_inicio, fila_titulo_mes, col_inicio + 3, nombre_mes, estilos["encabezado_meses"])
    hoja.set_row(fila_subencabezado, 28)
    hoja.write(fila_subencabezado, col_inicio,     "EGRESOS",            estilos["lista_unidades_txt"])
    hoja.set_column(col_inicio + 1, col_inicio + 1, 12)
    hoja.write(fila_subencabezado, col_inicio + 1, "INFECCIONES",        estilos["lista_unidades_txt"])
    hoja.write(fila_subencabezado, col_inicio + 2, "DÍAS ESTANCIA",      estilos["lista_unidades_txt"])
    hoja.write(fila_subencabezado, col_inicio + 3, "TASA POR 1000 DÍAS", estilos["lista_unidades_txt"])

    for i in range(cantidad_unidades):
        estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
        for offset in range(4):
            hoja.write(fila_subencabezado + 1 + i, col_inicio + offset, "", estilo)

    fila_delegacion = fila_subencabezado + cantidad_unidades
    for offset in range(4):
        hoja.write(fila_delegacion, col_inicio + offset, "", estilos["delegacion_txt"])


def _escribir_IASS01(libro, estilos):
    fila_titulo1       = 0
    fila_titulo2       = 1
    fila_titulo3       = 2
    fila_nota          = 3
    fila_encab_meses   = 4
    fila_subencabezado = 5
    col_fin_tabla      = 25

    hoja = libro.add_worksheet("IASS 01")
    hoja.set_column(0, 0, 27)
    hoja.set_default_row(20)
    hoja.freeze_panes(0, 1)
    libro.formats[0].set_font_name("Calibri")

    hoja.merge_range(fila_titulo1, 1, fila_titulo1, col_fin_tabla - 1, "INFORMES DE INFECCIONES NOSOCOMIALES", estilos["titulo"])
    hoja.merge_range(fila_titulo2, 1, fila_titulo2, col_fin_tabla - 1, "DELEGACIÓN GUANAJUATO",                estilos["titulo"])
    hoja.merge_range(fila_titulo3, 1, fila_titulo3, col_fin_tabla - 1, "INFORMES DE INFECCIONES NOSOCOMIALES", estilos["titulo"])
    hoja.merge_range(fila_nota,    1, fila_nota,    col_fin_tabla - 1,
        "NOTA:LA SUMATORIA INCLUYE UNICAMENTE LAS 8 UNIDADES QUE ENTRAN A EVALUACIÓN QUE SON LA 2, 3, 4, 10, 13, 21, 54, 58",
        estilos["titulo_tabla"])

    cantidad_unidades = len(UNIDADES_IASS)
    hoja.merge_range(fila_encab_meses, 0, fila_subencabezado, 0, "UNIDADES", estilos["unidades_txt"])
    _escribir_unidades(hoja, estilos, fila_subencabezado + 1, 0)

    col_inicio_mes = 1
    for mes in MESES:
        if mes == "JULIO":
            break
        _escribir_bloque_mesIASS_01(hoja, estilos, fila_encab_meses, fila_subencabezado, col_inicio_mes, cantidad_unidades, mes)
        col_inicio_mes += 4

    col_inicio_acumulado = 27
    for mes in MESES[1:]:
        if mes == "JULIO":
            break
        _escribir_bloque_mesIASS_01(hoja, estilos, fila_encab_meses, fila_subencabezado, col_inicio_acumulado, cantidad_unidades, mes)
        col_inicio_acumulado += 4

    fila_encab_meses   = fila_nota + 1 + cantidad_unidades + 4
    fila_subencabezado = fila_encab_meses + 1
    hoja.merge_range(fila_encab_meses, 0, fila_subencabezado, 0, "UNIDADES", estilos["unidades_txt"])
    _escribir_unidades(hoja, estilos, fila_subencabezado + 1, 0)

    col_inicio_mes = 1
    for mes in MESES[6:]:
        _escribir_bloque_mesIASS_01(hoja, estilos, fila_encab_meses, fila_subencabezado, col_inicio_mes, cantidad_unidades, mes)
        col_inicio_mes += 4

    col_inicio_acumulado = 27
    for mes in MESES[6:]:
        _escribir_bloque_mesIASS_01(hoja, estilos, fila_encab_meses, fila_subencabezado, col_inicio_acumulado, cantidad_unidades, mes)
        col_inicio_acumulado += 4

    fila_encab_meses   = fila_nota + 1 + cantidad_unidades + 4 + cantidad_unidades + 4
    fila_subencabezado = fila_encab_meses + 1
    hoja.merge_range(fila_encab_meses, 0, fila_subencabezado, 0, "UNIDADES", estilos["unidades_txt"])
    _escribir_unidades(hoja, estilos, fila_subencabezado + 1, 0)
    _escribir_bloque_mesIASS_01(hoja, estilos, fila_encab_meses, fila_subencabezado, 1, cantidad_unidades, "Anual")


def Excel_IASS01():
    salida  = io.BytesIO()
    libro   = xlsxwriter.Workbook(salida)
    libro.set_properties({'author': 'Web CIAE'})
    estilos = Estilos_IASS01(libro)
    _escribir_IASS01(libro, estilos)
    libro.close()
    salida.seek(0)
    return salida


_COLOR_VERDE  = "#0B5445"
_COLOR_DORADO = "#9A7026"
_COLOR_ROJO   = "#7E0808"


def Estilos_IASS01(libro):
    _cap = {"font_size": 9, "align": "center", "valign": "vcenter",
            "bold": True, "border": 1, "border_color": "#A6A6A6", "num_format": "0.00"}
    return {
        "titulo":               libro.add_format({"font_size": 9, "align": "center"}),
        "titulo_tabla":         libro.add_format({"font_size": 16, "align": "center", "bold": True}),
        "unidades_txt":         libro.add_format({"font_size": 11, "align": "center", "bold": True, "valign": "vcenter",
                                                   "bg_color": "#F2F2F2", "border_color": "#A6A6A6", "border": 1}),
        "lista_unidades_txt":   libro.add_format({"font_size": 9, "bg_color": "#F2F2F2", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "center", "text_wrap": True}),
        "delegacion_txt":       libro.add_format({"font_size": 9, "bg_color": "#808080", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "center", "text_wrap": True,
                                                   "bold": True, "font_color": "#FFFFFF"}),
        "encabezado_meses":     libro.add_format({"font_size": 11, "bg_color": "#F2F2F2", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "center", "bold": True,
                                                   "font_color": "#3D3D3D"}),
        "color_celda1":         libro.add_format({"font_size": 9, "bg_color": "#f9f9f9", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "center",
                                                   "text_wrap": True, "bold": True, "font_color": "#000000"}),
        "color_celda2":         libro.add_format({"font_size": 9, "bg_color": "#FFFFFF", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "center",
                                                   "text_wrap": True, "bold": True, "font_color": "#000000"}),
        "titulo_iass06":        libro.add_format({"font_size": 11, "bold": True, "align": "center", "valign": "vcenter",
                                                   "border": 1, "border_color": "#A6A6A6"}),
        "header_grupo2":        libro.add_format({"font_size": 9, "bg_color": "#808080", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "left", "bold": True,
                                                   "font_color": "#FFFFFF", "text_wrap": True}),
        "Verde":    libro.add_format({**_cap, "bg_color": _COLOR_VERDE,  "font_color": "white"}),
        "Amarillo": libro.add_format({**_cap, "bg_color": _COLOR_DORADO, "font_color": "white"}),
        "Rojo":     libro.add_format({**_cap, "bg_color": _COLOR_ROJO,   "font_color": "white"}),
        "sin_datos":libro.add_format({**_cap, "bg_color": "#CCCCCC",     "font_color": "black"}),
    }


_IASS_UCI_CONFIG = {
    num: {
        "codigo":   _CFG[f"IASS {num}"]["Titulo2"],
        "titulo":   _CFG[f"IASS {num}"]["Titulo1"],
        "hoja":     f"IASS {num}",
        "sub_col1": _CFG[f"IASS {num}"]["subT1"],
        "sub_col2": _CFG[f"IASS {num}"]["subT2"],
        "alto_sub": 48 if num == "04" else 30,
    }
    for num in ("02", "03", "04", "05", "06")
}


def _escribir_tabla_iass_uci(hoja, estilos, fila_inicio, etiqueta_seccion, cfg):
    COLS_MES      = 3
    TOTAL_COLS    = 1 + 12 * COLS_MES
    cant_unidades = len(UNIDADES_UCI)

    fila_seccion = fila_inicio
    fila_mes     = fila_seccion + 1
    fila_sub     = fila_mes + 1
    fila_hosp1   = fila_sub + 1

    hoja.write(fila_seccion, 0, cfg["codigo"], estilos["lista_unidades_txt"])
    hoja.merge_range(fila_seccion, 1, fila_seccion, TOTAL_COLS - 1, etiqueta_seccion, estilos["titulo_iass06"])
    hoja.merge_range(fila_mes, 0, fila_sub, 0, "UNIDADES", estilos["unidades_txt"])
    hoja.set_row(fila_sub, cfg["alto_sub"])

    col = 1
    for mes in MESES:
        hoja.merge_range(fila_mes, col, fila_mes, col + 2, mes, estilos["encabezado_meses"])
        hoja.write(fila_sub, col,     cfg["sub_col1"], estilos["lista_unidades_txt"])
        hoja.write(fila_sub, col + 1, cfg["sub_col2"], estilos["lista_unidades_txt"])
        hoja.write(fila_sub, col + 2, "Tasa",          estilos["lista_unidades_txt"])
        col += COLS_MES

    for i, unidad in enumerate(UNIDADES_UCI):
        fila   = fila_hosp1 + i
        estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
        hoja.write(fila, 0, unidad, estilos["lista_unidades_txt"])
        for c in range(1, TOTAL_COLS):
            hoja.write(fila, c, "", estilo)

    return fila_hosp1 + cant_unidades - 1


def _escribir_IASS_UCI(libro, estilos, numero: str):
    cfg        = _IASS_UCI_CONFIG[numero]
    COLS_MES   = 3
    TOTAL_COLS = 1 + 12 * COLS_MES

    hoja = libro.add_worksheet(cfg["hoja"])
    hoja.set_column(0, 0, 32)
    hoja.freeze_panes(0, 1)
    hoja.set_column(1, TOTAL_COLS - 1, 9)
    for m in range(12):
        hoja.set_column(2 + m * COLS_MES, 2 + m * COLS_MES, 14)
    hoja.set_default_row(20)
    libro.formats[0].set_font_name("Calibri")

    hoja.merge_range(0, 0, 0, TOTAL_COLS - 1, cfg["titulo"], estilos["titulo_iass06"])

    fila_fin_mensual   = _escribir_tabla_iass_uci(hoja, estilos, 1,                    "MENSUAL",           cfg)
    fila_fin_acumulado = _escribir_tabla_iass_uci(hoja, estilos, fila_fin_mensual + 2, "MENSUAL ACUMULADO", cfg)

    fila_delega = fila_fin_acumulado + 1
    hoja.write(fila_delega, 0, "DELEGACION", estilos["delegacion_txt"])
    for c in range(1, TOTAL_COLS):
        hoja.write(fila_delega, c, "", estilos["delegacion_txt"])


def _Excel_IASS_UCI(numero: str) -> io.BytesIO:
    salida  = io.BytesIO()
    libro   = xlsxwriter.Workbook(salida)
    libro.set_properties({'author': 'Web CIAE'})
    estilos = Estilos_IASS01(libro)
    _escribir_IASS_UCI(libro, estilos, numero)
    libro.close()
    salida.seek(0)
    return salida


def Excel_IASS02(): return _Excel_IASS_UCI("02")
def Excel_IASS03(): return _Excel_IASS_UCI("03")
def Excel_IASS04(): return _Excel_IASS_UCI("04")
def Excel_IASS05(): return _Excel_IASS_UCI("05")
def Excel_IASS06(): return _Excel_IASS_UCI("06")


def _estilo_tasa(estilos, color):
    return estilos.get(color, estilos["sin_datos"])


def _escribir_acumulado_IASS01(hoja, estilos, all_months):
    N = len(UNIDADES_IASS)

    for mes_target in range(2, 13):
        if mes_target not in all_months:
            continue

        if mes_target <= 6:
            fila_inicio = 6
            col_base    = 27 + (mes_target - 2) * 4
        else:
            fila_inicio = N + 10
            col_base    = 27 + (mes_target - 7) * 4

        sum_del_n = 0
        sum_del_d = 0

        for i, unidad in enumerate(UNIDADES_IASS):
            if unidad == "DELEGACION":
                continue

            acum_num = 0
            acum_den = 0
            tiene    = False
            for m in range(1, mes_target + 1):
                v = all_months.get(m, {}).get(unidad)
                if isinstance(v, dict):
                    acum_num += v.get("numerador") or 0
                    acum_den += v.get("denominador") or 0
                    tiene = True

            if not tiene:
                continue

            fila   = fila_inicio + i
            estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
            tasa   = round((acum_num / acum_den) * 1000, 2) if acum_den else 0
            color  = _color_tasa_01(tasa, unidad)

            hoja.write(fila, col_base + 1, acum_num, estilo)
            hoja.write(fila, col_base + 2, acum_den, estilo)
            hoja.write(fila, col_base + 3, tasa, _estilo_tasa(estilos, color))

            sum_del_n += acum_num
            sum_del_d += acum_den

        fila_del  = fila_inicio + N - 1
        tasa_del  = round((sum_del_n / sum_del_d) * 1000, 2) if sum_del_d else 0
        color_del = _color_tasa_01(tasa_del, "DELEGACION")
        est_del   = estilos["delegacion_txt"]
        hoja.write(fila_del, col_base + 1, sum_del_n, est_del)
        hoja.write(fila_del, col_base + 2, sum_del_d, est_del)
        hoja.write(fila_del, col_base + 3, tasa_del,  _estilo_tasa(estilos, color_del))


def _escribir_acumulado_IASS_uci(hoja, estilos, all_months, indicador):
    N         = len(UNIDADES_UCI)
    fila_ini  = N + 8
    fila_del  = 2 * N + 8

    sem       = _CFG[indicador].get("Semaforo", {})
    tasa_mult = sem.get("Tasa", 1000)

    for mes_target in range(1, 13):
        if mes_target not in all_months:
            continue

        col_base  = 1 + (mes_target - 1) * 3
        sum_del_n = 0
        sum_del_d = 0

        for i, unidad in enumerate(UNIDADES_UCI):
            acum_num = 0
            acum_den = 0
            tiene    = False
            for m in range(1, mes_target + 1):
                v = all_months.get(m, {}).get(unidad)
                if isinstance(v, dict):
                    acum_num += v.get("numerador") or 0
                    acum_den += v.get("denominador") or 0
                    tiene = True

            if not tiene:
                continue

            fila   = fila_ini + i
            estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
            tasa   = round((acum_num / acum_den) * tasa_mult, 2) if acum_den else 0
            color  = _color_tasa_uci(tasa, indicador)

            hoja.write(fila, col_base,     acum_num, estilo)
            hoja.write(fila, col_base + 1, acum_den, estilo)
            hoja.write(fila, col_base + 2, tasa,     _estilo_tasa(estilos, color))

            sum_del_n += acum_num
            sum_del_d += acum_den

        tasa_del  = round((sum_del_n / sum_del_d) * tasa_mult, 2) if sum_del_d else 0
        color_del = _color_tasa_uci(tasa_del, indicador)
        est_del   = estilos["delegacion_txt"]
        hoja.write(fila_del, col_base,     sum_del_n, est_del)
        hoja.write(fila_del, col_base + 1, sum_del_d, est_del)
        hoja.write(fila_del, col_base + 2, tasa_del,  _estilo_tasa(estilos, color_del))


def _escribir_datos_IASS_uci(hoja, estilos, mes_num, datos):
    col_base = 1 + (mes_num - 1) * 3
    for i, unidad in enumerate(UNIDADES_UCI):
        v      = datos.get(unidad)
        fila   = 4 + i
        estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
        if not isinstance(v, dict):
            hoja.write(fila, col_base,     0, estilo)
            hoja.write(fila, col_base + 1, 0, estilo)
            hoja.write(fila, col_base + 2, 0, _estilo_tasa(estilos, "Rojo"))
            continue
        num  = v.get("numerador")
        den  = v.get("denominador")
        tasa = v.get("tasa")
        hoja.write(fila, col_base,     num  if num  is not None else 0, estilo)
        hoja.write(fila, col_base + 1, den  if den  is not None else 0, estilo)
        hoja.write(fila, col_base + 2, tasa if tasa is not None else 0, _estilo_tasa(estilos, v.get("color")))


def _escribir_datos_IASS01(hoja, estilos, mes_num, datos):
    N = len(UNIDADES_IASS)
    if mes_num <= 6:
        fila_inicio = 6
        col_base    = 1 + (mes_num - 1) * 4
    else:
        fila_inicio = N + 10
        col_base    = 1 + (mes_num - 7) * 4

    unidades_datos = {k: v for k, v in datos.items() if isinstance(v, dict) and k not in ("Delegación", "DELEGACION")}
    sum_num        = sum((v.get("numerador") or 0) for v in unidades_datos.values())
    sum_den        = sum((v.get("denominador") or 0) for v in unidades_datos.values())
    tasa_deleg     = round((sum_num / sum_den) * 1000, 2) if sum_den else 0

    for i, unidad in enumerate(UNIDADES_IASS):
        fila = fila_inicio + i

        if unidad == "DELEGACION":
            est         = estilos["delegacion_txt"]
            color_deleg = datos.get("Delegación", {}).get("color") if isinstance(datos.get("Delegación"), dict) else _color_tasa_01(tasa_deleg, "DELEGACION")
            hoja.write(fila, col_base + 1, sum_num,    est)
            hoja.write(fila, col_base + 2, sum_den,    est)
            hoja.write(fila, col_base + 3, tasa_deleg, _estilo_tasa(estilos, color_deleg))
            continue

        v = datos.get(unidad)
        if not isinstance(v, dict):
            continue
        estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
        num  = v.get("numerador")
        den  = v.get("denominador")
        tasa = v.get("tasa")
        hoja.write(fila, col_base + 1, num  if num  is not None else 0, estilo)
        hoja.write(fila, col_base + 2, den  if den  is not None else 0, estilo)
        hoja.write(fila, col_base + 3, tasa if tasa is not None else 0, _estilo_tasa(estilos, v.get("color")))


def Excel_IASS_Completo(anio: str, mes: str, datos: dict) -> io.BytesIO:
    mes_num      = int(mes)
    ruta_archivo = RUTA_BASE_IASS / str(anio) / f"IASS_{anio}.xlsx"
    historicos   = _leer_historicos_IASS(anio, mes_num)

    salida  = io.BytesIO()
    libro   = xlsxwriter.Workbook(salida)
    libro.set_properties({'author': 'Web CIAE'})
    estilos = Estilos_IASS01(libro)

    _escribir_IASS01(libro, estilos)
    for num in ("02", "03", "04", "05", "06"):
        _escribir_IASS_UCI(libro, estilos, num)

    hoja01 = libro.get_worksheet_by_name("IASS 01")
    if hoja01:
        all_months_01 = dict(historicos.get("IASS 01", {}))
        for m, mes_datos in all_months_01.items():
            _escribir_datos_IASS01(hoja01, estilos, m, mes_datos)
        if datos.get("IASS 01"):
            all_months_01[mes_num] = datos["IASS 01"]
            _escribir_datos_IASS01(hoja01, estilos, mes_num, datos["IASS 01"])
        _escribir_acumulado_IASS01(hoja01, estilos, all_months_01)

    for num in ("02", "03", "04", "05", "06"):
        key  = f"IASS {num}"
        hoja = libro.get_worksheet_by_name(key)
        if not hoja:
            continue

        all_months = dict(historicos.get(key, {}))
        if datos.get(key):
            all_months[mes_num] = datos[key]

        for m, mes_datos in all_months.items():
            _escribir_datos_IASS_uci(hoja, estilos, m, mes_datos)

        _escribir_acumulado_IASS_uci(hoja, estilos, all_months, key)

    libro.close()
    salida.seek(0)

    ruta_archivo.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(ruta_archivo, "wb") as f:
            f.write(salida.read())
        salida.seek(0)
        print(f"[IASS] Guardado: {ruta_archivo}")
    except PermissionError:
        salida.seek(0)
        raise PermissionError(f"El archivo IASS_{anio}.xlsx está abierto en Excel. Ciérralo e intenta de nuevo.")

    return salida

"""
Módulo  : reportes_controller.py
Carpeta : reportes/controllers/
Qué hace: Endpoints para generar y descargar reportes Excel (FTP, IN_ASS, categoría).
Usado en: main.py (prefix /reportes)
"""
import asyncio
import base64
import io
import json
import traceback
from typing import Optional, List

import openpyxl
import xlsxwriter
from fastapi import APIRouter, Query, Depends, HTTPException, status, UploadFile, File, Form, Request

from configs.response import ApiResponse
from auth.services.jwt_utils import solo_roles
from indicadores_FTP.services.FTP_service import obtenerInformacionIndicador, consultarTodosIndicadores
from indicadores_IN_ASS.services.procesar_service import procesar_in_ass as ProcesarInAss
from reportes.services.reporte_final import ExcelReporteFinal
from reportes.services.reporte_categoria import preparar_datos_indicador, escribir_hoja_indicador
from reportes.services.generar_excel import obtener_estilos_excel, _calcular_color
from reportes.services.generar_in_ass import (
    Excel_In_Ass01, Excel_In_Ass02, Excel_In_Ass03,
    Excel_In_Ass04, Excel_In_Ass05, Excel_In_Ass06,
    RUTA_BASE_IN_ASS, UNIDADES_UCI, UNIDADES_IASS,
    _color_tasa_uci, _color_tasa_01
) 
from reportes.services.datos_json_service import (
    meses_con_datos as ftp_meses_con_datos,
    leer_datos_indicador,
    leer_semana_indicador,
    MESES_NOMBRES as FTP_MESES_NOMBRES,
)

router = APIRouter()

ROLES_FTP_FULL  = ("admin", "trabajador_ftp")
ROLES_TODOS     = ("admin", "trabajador_ftp", "trabajador_IAAS", "visitante")
ROLES_FTP_GRAF  = ROLES_TODOS
ROLES_INASS     = ("admin", "trabajador_IAAS")
ROLES_INASS_GRAF = ROLES_TODOS


# ─── /meses-generados ────────────────────────────────────────────────────────

@router.get("/meses-generados")
async def meses_generados(
    indicador: str = Query(...),
    ano:       str = Query(...),
    payload:   dict = Depends(solo_roles(*ROLES_FTP_FULL))
):
    meses = ftp_meses_con_datos(indicador, ano)
    return ApiResponse(success=True, message="Meses con datos obtenidos", data={"meses": meses})


# ─── /Indicadores ─────────────────────────────────────────────────────────────

@router.get("/Indicadores")
async def reporte(
    indicador: str = Query(...),
    ano:       str = Query(...),
    mes:       str = Query(...),
    semana:    Optional[str] = Query(None),
    payload:   dict = Depends(solo_roles(*ROLES_FTP_FULL))
):
    if semana and payload.get("rol") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Sin acceso a reportes semanales")

    resultado = ExcelReporteFinal(indicador, ano, mes, semana)

    if resultado["status"] == "success":
        archivo_buffer = resultado["stream"]
        excel_b64      = base64.b64encode(archivo_buffer.getvalue()).decode("utf-8")
        return ApiResponse(success=True, message=resultado.get("mensaje", "Reporte generado"), data={
            "archivo_b64":    excel_b64,
            "nombre_archivo": resultado["nombre_archivo"],
            "datos_grafica":  resultado.get("graficar", {}),
            "restricciones":  resultado.get("restricciones", {}),
        })
    else:
        return ApiResponse(success=False, message=resultado.get("mensaje", "Error desconocido"), data={
            "restricciones": resultado.get("restricciones"),
        })





# ─── /FTP/datos-grafica ───────────────────────────────────────────────────────

@router.get("/FTP/datos-grafica")
async def ftp_datos_grafica(
    indicador: str = Query(...),
    anio:      str = Query(...),
    payload:   dict = Depends(solo_roles(*ROLES_FTP_GRAF))
):
    try:
        info     = obtenerInformacionIndicador(indicador)
        semaforo = info.get("semaforo", {})
        datos_json = leer_datos_indicador(indicador, anio)

        if not datos_json:
            return ApiResponse(success=True, message="Sin historial", data={"unidades": [], "meses_con_datos": [], "datos": {}})

        datos     = {}
        meses_set = set()
        unidades_set = []

        for mes_nombre, unidades_mes in datos_json.get("MESES", {}).items():
            if mes_nombre not in FTP_MESES_NOMBRES:
                continue
            idx_mes = FTP_MESES_NOMBRES.index(mes_nombre)
            mes_str = str(idx_mes + 1).zfill(2)

            for unidad, vals in unidades_mes.items():
                num = vals.get("numerador")
                den = vals.get("denominador")
                pct = vals.get("%")
                if num is None and pct is None:
                    continue
                meses_set.add(mes_str)
                color_guardado = vals.get("color")
                color = (
                    _calcular_color(pct, idx_mes, semaforo)
                    if not color_guardado or color_guardado == "Gris"
                    else color_guardado
                )
                datos.setdefault(unidad, []).append({
                    "mes":         mes_str,
                    "tasa":        float(pct) if pct is not None else 0,
                    "numerador":   float(num) if num is not None else 0,
                    "denominador": float(den) if den is not None else 0,
                    "color":       color,
                })
                if unidad not in unidades_set:
                    unidades_set.append(unidad)

        # Ordenar puntos de cada unidad por mes
        for unidad in datos:
            datos[unidad].sort(key=lambda x: x["mes"])

        # Mezclar datos semanales para meses sin reporte definitivo
        meses_definitivos = set(meses_set)
        semanal_json      = leer_semana_indicador(indicador, anio)
        for mes_nombre, mes_data in semanal_json.get("MESES", {}).items():
            if mes_nombre not in FTP_MESES_NOMBRES:
                continue
            idx_mes = FTP_MESES_NOMBRES.index(mes_nombre)
            mes_str = str(idx_mes + 1).zfill(2)
            if mes_str in meses_definitivos:
                continue  # ya existe definitivo, no usar semanal
            semana_num = mes_data.get("semana", 1)
            meses_set.add(mes_str)
            for unidad, vals in mes_data.items():
                if unidad == "semana" or not isinstance(vals, dict):
                    continue
                pct = vals.get("%")
                num = vals.get("numerador")
                den = vals.get("denominador")
                col = vals.get("color", "Gris")
                if unidad not in unidades_set:
                    unidades_set.append(unidad)
                datos.setdefault(unidad, []).append({
                    "mes":         mes_str,
                    "tasa":        float(pct) if pct is not None else 0,
                    "numerador":   float(num) if num is not None else 0,
                    "denominador": float(den) if den is not None else 0,
                    "color":       col,
                    "semana":      semana_num,
                })
        for unidad in datos:
            datos[unidad].sort(key=lambda x: x["mes"])

        return ApiResponse(success=True, message="Datos de gráfica obtenidos", data={
            "unidades":        unidades_set,
            "meses_con_datos": sorted(meses_set),
            "datos":           datos,
        })
    except Exception as e:
        print(f"[FTP datos-grafica] {indicador}: {e}")
        return ApiResponse(success=False, message=str(e), data={"unidades": [], "meses_con_datos": [], "datos": {}})





# ─── /InAss/* ─────────────────────────────────────────────────────────────────

@router.get("/ASS_PRUEBA")
async def IASS(payload: dict = Depends(solo_roles(*ROLES_INASS))):
    archivo   = Excel_In_Ass01()
    excel_b64 = base64.b64encode(archivo.getvalue()).decode("utf-8")
    return ApiResponse(success=True, message="IN_ASS 01 generado", data={"archivo_b64": excel_b64, "nombre_archivo": "IN_ASS_01.xlsx"})


@router.get("/ASS_02")
async def IASS_02(payload: dict = Depends(solo_roles(*ROLES_INASS))):
    excel_b64 = base64.b64encode(Excel_In_Ass02().getvalue()).decode("utf-8")
    return ApiResponse(success=True, message="IN_ASS 02 generado", data={"archivo_b64": excel_b64, "nombre_archivo": "IN_ASS_02.xlsx"})


@router.get("/ASS_03")
async def IASS_03(payload: dict = Depends(solo_roles(*ROLES_INASS))):
    excel_b64 = base64.b64encode(Excel_In_Ass03().getvalue()).decode("utf-8")
    return ApiResponse(success=True, message="IN_ASS 03 generado", data={"archivo_b64": excel_b64, "nombre_archivo": "IN_ASS_03.xlsx"})


@router.get("/ASS_04")
async def IASS_04(payload: dict = Depends(solo_roles(*ROLES_INASS))):
    excel_b64 = base64.b64encode(Excel_In_Ass04().getvalue()).decode("utf-8")
    return ApiResponse(success=True, message="IN_ASS 04 generado", data={"archivo_b64": excel_b64, "nombre_archivo": "IN_ASS_04.xlsx"})


@router.get("/ASS_05")
async def IASS_05(payload: dict = Depends(solo_roles(*ROLES_INASS))):
    excel_b64 = base64.b64encode(Excel_In_Ass05().getvalue()).decode("utf-8")
    return ApiResponse(success=True, message="IN_ASS 05 generado", data={"archivo_b64": excel_b64, "nombre_archivo": "IN_ASS_05.xlsx"})


@router.get("/ASS_06")
async def IASS_06(payload: dict = Depends(solo_roles(*ROLES_INASS))):
    excel_b64 = base64.b64encode(Excel_In_Ass06().getvalue()).decode("utf-8")
    return ApiResponse(success=True, message="IN_ASS 06 generado", data={"archivo_b64": excel_b64, "nombre_archivo": "IN_ASS_06.xlsx"})


@router.get("/InAss/meses-guardados")
async def inass_meses_guardados(
    anio:    str = Query(...),
    payload: dict = Depends(solo_roles(*ROLES_INASS_GRAF))
):
    ruta = RUTA_BASE_IN_ASS / str(anio) / f"IN_ASS_{anio}.xlsx"
    if not ruta.exists():
        return ApiResponse(success=True, message="Sin archivo guardado", data={"meses": []})
    try:
        wb    = openpyxl.load_workbook(ruta, data_only=True)
        ws    = wb["IN_ASS 02"]
        meses = []
        for m in range(1, 13):
            col   = 2 + (m - 1) * 3
            tiene = any(
                isinstance(ws.cell(row=r, column=col).value, (int, float))
                for r in range(5, 35)
            )
            if tiene:
                meses.append(str(m).zfill(2))
        wb.close()
        return ApiResponse(success=True, message="Meses guardados obtenidos", data={"meses": meses})
    except Exception as e:
        print(f"[IN_ASS meses-guardados] {e}")
        return ApiResponse(success=False, message=str(e), data={"meses": []})


@router.get("/InAss/descargar")
async def inass_descargar(
    anio:    str = Query(...),
    payload: dict = Depends(solo_roles(*ROLES_INASS))
):
    ruta = RUTA_BASE_IN_ASS / str(anio) / f"IN_ASS_{anio}.xlsx"
    if not ruta.exists():
        raise HTTPException(status_code=404, detail="No existe archivo guardado para ese año")
    with open(ruta, "rb") as f:
        contenido = f.read()
    return ApiResponse(success=True, message=f"IN_ASS_{anio}.xlsx", data={
        "archivo_b64":    base64.b64encode(contenido).decode("utf-8"),
        "nombre_archivo": f"IN_ASS_{anio}.xlsx"
    })




@router.get("/InAss/datos-grafica")
async def inass_datos_grafica(
    anio:    str = Query(...),
    payload: dict = Depends(solo_roles(*ROLES_INASS_GRAF))
):
    ruta = RUTA_BASE_IN_ASS / str(anio) / f"IN_ASS_{anio}.xlsx"
    if not ruta.exists():
        return ApiResponse(success=True, message="Sin datos guardados", data={"unidades": [], "meses_con_datos": [], "datos": {}})
    try:
        wb        = openpyxl.load_workbook(ruta, data_only=True)
        datos     = {}
        meses_set = set()
        N         = len(UNIDADES_IASS)

        try:
            ws01     = wb["IN_ASS 01"]
            ind_data = {}
            for m in range(1, 13):
                mes_str = str(m).zfill(2)
                if m <= 6:
                    fila_ini = 7
                    col_base = 1 + (m - 1) * 4
                else:
                    fila_ini = N + 11
                    col_base = 1 + (m - 7) * 4
                col_num  = col_base + 2
                col_den  = col_base + 3
                col_tasa = col_base + 4
                mes_vals = {}
                for i, unidad in enumerate(UNIDADES_IASS):
                    if unidad == "DELEGACION":
                        continue
                    row = fila_ini + i
                    n_v = ws01.cell(row=row, column=col_num).value
                    d_v = ws01.cell(row=row, column=col_den).value
                    t_v = ws01.cell(row=row, column=col_tasa).value
                    if isinstance(n_v, (int, float)) or isinstance(t_v, (int, float)):
                        mes_vals[unidad] = {
                            "mes":         mes_str,
                            "tasa":        t_v if isinstance(t_v, (int, float)) else 0,
                            "numerador":   n_v if isinstance(n_v, (int, float)) else 0,
                            "denominador": d_v if isinstance(d_v, (int, float)) else 0,
                            "color":       _color_tasa_01(t_v, unidad),
                        }
                if mes_vals:
                    meses_set.add(mes_str)
                    for unidad, registro in mes_vals.items():
                        ind_data.setdefault(unidad, []).append(registro)
            datos["IN_ASS 01"] = ind_data
        except Exception as e:
            print(f"[IN_ASS datos-grafica] IN_ASS 01: {e}")

        for num in ("02", "03", "04", "05", "06"):
            key = f"IN_ASS {num}"
            try:
                ws       = wb[key]
                ind_data = {}
                for m in range(1, 13):
                    col_num  = 2 + (m - 1) * 3
                    col_den  = 3 + (m - 1) * 3
                    col_tasa = 4 + (m - 1) * 3
                    mes_str  = str(m).zfill(2)
                    mes_vals = {}
                    for i, unidad in enumerate(UNIDADES_UCI):
                        row = 5 + i
                        n_v = ws.cell(row=row, column=col_num).value
                        d_v = ws.cell(row=row, column=col_den).value
                        t_v = ws.cell(row=row, column=col_tasa).value
                        if isinstance(n_v, (int, float)) or isinstance(t_v, (int, float)):
                            mes_vals[unidad] = {
                                "mes":         mes_str,
                                "tasa":        t_v if isinstance(t_v, (int, float)) else 0,
                                "numerador":   n_v if isinstance(n_v, (int, float)) else 0,
                                "denominador": d_v if isinstance(d_v, (int, float)) else 0,
                                "color":       _color_tasa_uci(t_v, key),
                            }
                    if mes_vals:
                        meses_set.add(mes_str)
                        for unidad, registro in mes_vals.items():
                            ind_data.setdefault(unidad, []).append(registro)
                datos[key] = ind_data
            except Exception as e:
                print(f"[IN_ASS datos-grafica] {key}: {e}")

        wb.close()
        return ApiResponse(success=True, message="Datos de gráfica IN_ASS obtenidos", data={
            "unidades":        list(UNIDADES_UCI),
            "meses_con_datos": sorted(meses_set),
            "datos":           datos
        })
    except Exception as e:
        print(f"[IN_ASS datos-grafica] {e}")
        return ApiResponse(success=False, message=str(e), data={"unidades": [], "meses_con_datos": [], "datos": {}})





@router.post("/InAss/Generar")
async def RecibirDatosInAss(
    anio:                       str                  = Form(...),
    mes:                        str                  = Form(...),
    numerador:                  List[UploadFile]     = File(default=[]),
    denominador:                Optional[str]        = Form(None),
    excel_denominador_inass_01: Optional[UploadFile] = File(None),
    payload:                    dict                 = Depends(solo_roles(*ROLES_INASS))
):
    try:
        numerador_bytes            = {f.filename: await f.read() for f in numerador}
        excel_denominador_01_bytes = await excel_denominador_inass_01.read() if excel_denominador_inass_01 else None
        denominador_dict           = json.loads(denominador) if denominador else {}

        resultado = ProcesarInAss(
            anio                       = anio,
            mes                        = mes,
            numerador                  = numerador_bytes,
            denominador                = denominador_dict,
            excel_denominador_inass_01 = excel_denominador_01_bytes
        )
        return ApiResponse(success=True, message=resultado["mensaje"], data={
            "archivo_b64":    resultado["archivo_b64"],
            "nombre_archivo": resultado["nombre_archivo"]
        })
    except ValueError as e:
        msg = str(e)
        try:
            detail = json.loads(msg)
        except Exception:
            detail = msg
        raise HTTPException(status_code=400, detail=detail)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error al recibir datos IN_ASS: {str(e)}")




# ─── /recalcular-poblacion ────────────────────────────────────────────────────

@router.post("/recalcular-poblacion")
async def recalcular_poblacion(request: Request, payload: dict = Depends(solo_roles(*ROLES_FTP_FULL))):
    """
    Recalcula denominador y resultado de los indicadores que usan POBLACION.json,
    leyendo el numerador del Excel histórico ya guardado. NO accede al FTP,
    por lo que no hay riesgo de perder datos históricos que ya no estén en el servidor.
    """
    from reportes.services.recalcular_poblacion_service import actualizar_historico_con_nueva_poblacion

    body = await request.json()
    ano  = str(body.get("ano", "2026"))

    todos        = consultarTodosIndicadores()
    recalculados = []
    errores      = []

    for categoria, cat_data in todos.items():
        for indicador in cat_data.get("indicadores", []):
            info     = obtenerInformacionIndicador(indicador)
            reportes = info.get("reporte", {}) if isinstance(info, dict) else {}

            usa_poblacion = any(
                isinstance(v, dict) and v.get("modo") == "JSON_POBLACION"
                for v in reportes.values()
            )
            if not usa_poblacion:
                continue

            try:
                ok, detalle, n_meses = actualizar_historico_con_nueva_poblacion(indicador, ano, info)
                if ok:
                    recalculados.append({"indicador": indicador, "meses": n_meses, "detalle": detalle})
                else:
                    errores.append(f"{indicador}: {detalle}")
            except Exception as e:
                errores.append(f"{indicador}: {str(e)}")

    return ApiResponse(success=True, message="Recálculo completado", data={
        "total":        sum(r["meses"] for r in recalculados),
        "recalculados": recalculados,
        "errores":      errores,
    })




# ─── /generar-categoria ───────────────────────────────────────────────────────

@router.post("/generar-categoria")
async def generar_categoria(request: Request, payload: dict = Depends(solo_roles(*ROLES_FTP_FULL))):
    body      = await request.json()
    categoria = body.get("categoria", "").strip()
    ano       = body.get("ano", "").strip()
    mes       = body.get("mes", "").strip()
    semana    = body.get("semana")

    if not all([categoria, ano, mes]):
        raise HTTPException(status_code=400, detail="Faltan parámetros: categoria, ano, mes")

    todos    = consultarTodosIndicadores()
    cat_data = todos.get(categoria)
    if not cat_data:
        raise HTTPException(status_code=404, detail=f"Categoría '{categoria}' no encontrada")

    indicadores = cat_data.get("indicadores", [])
    if not indicadores:
        raise HTTPException(status_code=404, detail="No hay indicadores habilitados en esta categoría")

    es_semana = bool(semana and str(semana).strip() not in ("", "None", "none"))
    loop      = asyncio.get_running_loop()

    pares = []
    for ind in indicadores:
        resultado = await loop.run_in_executor(None, preparar_datos_indicador, ind, ano, mes, semana)
        pares.append((ind, resultado))

    output = io.BytesIO()
    wb     = xlsxwriter.Workbook(output)
    fmt    = obtener_estilos_excel(wb)

    completados   = []
    errores       = {}
    restricciones = {}

    for indicador, resultado in pares:
        if resultado["status"] != "success":
            errores[indicador] = resultado.get("mensaje", "Error desconocido")
            continue
        try:
            escribir_hoja_indicador(
                wb, fmt, indicador,
                resultado["diccionarioPrevio"],
                resultado["metadata"],
                ano, mes, semana, es_semana
            )
            completados.append(indicador)
            log_ftp = resultado.get("errores") or {}
            if log_ftp:
                errores[indicador] = log_ftp
                for tipo, val in log_ftp.items():
                    if tipo not in restricciones:
                        restricciones[tipo] = {
                            "nombreError":      val["nombreError"],
                            "descripcionError": val["descripcionError"],
                            "unidades":         {},
                        }
                    for unidad, paths in val.get("unidades", {}).items():
                        clave = f"{indicador} / {unidad}"
                        restricciones[tipo]["unidades"][clave] = paths
        except Exception as exc:
            errores[indicador] = str(exc)

    wb.close()

    if not completados:
        return ApiResponse(success=False, message="Ningún indicador pudo generarse", data={"errores": errores})

    output.seek(0)
    excel_b64 = base64.b64encode(output.getvalue()).decode("utf-8")
    mes_fmt   = str(mes).zfill(2)
    nombre    = f"{categoria}_{ano}_{mes_fmt}_S{semana}.xlsx" if es_semana else f"{categoria}_{ano}_{mes_fmt}.xlsx"

    return ApiResponse(success=True, message="Categoría generada", data={
        "archivo_b64":    excel_b64,
        "nombre_archivo": nombre,
        "completados":    completados,
        "errores":        errores,
        "restricciones":  restricciones,
    })

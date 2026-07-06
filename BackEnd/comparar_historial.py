"""
Script de comparacion: muestra diferencias entre Excel historico y JSON migrado.
  python comparar_historial.py
"""
import json
from pathlib import Path
import openpyxl

BASE_DIR       = Path(__file__).resolve().parent
DATA_DIR       = BASE_DIR.parent / "Data"
RUTA_HISTORIAL = DATA_DIR / "FTP" / "Historial_xlsx"
RUTA_DATA_FTP  = DATA_DIR / "FTP" / "Data_FTP"

MESES_NOMBRES = [
    "ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
    "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"
]

EXCLUIR = {"IASS", "Px_Referidos"}

def leer_excel(ruta_excel):
    """Lee el Excel y devuelve {mes_nombre: {unidad: {num, den, pct}}}"""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    unidades = []
    fila = 11
    while True:
        val = ws.cell(row=fila, column=1).value
        if val is None:
            break
        nombre = str(val).strip()
        if nombre.upper() != "TOTAL":
            unidades.append(nombre)
        fila += 1

    resultado = {}
    for idx_mes in range(12):
        col_num = (idx_mes * 3) + 2
        datos_mes = {}
        for idx_u, unidad in enumerate(unidades):
            num = ws.cell(row=11 + idx_u, column=col_num).value
            den = ws.cell(row=11 + idx_u, column=col_num + 1).value
            pct = ws.cell(row=11 + idx_u, column=col_num + 2).value
            if num is not None or den is not None or pct is not None:
                datos_mes[unidad] = (num, den, pct)
        if datos_mes:
            resultado[MESES_NOMBRES[idx_mes]] = datos_mes

    wb.close()
    return resultado

def leer_json(ruta_json):
    """Lee el JSON y devuelve {mes_nombre: {unidad: (num, den, pct)}}"""
    if not ruta_json.exists():
        return {}
    with open(ruta_json, encoding="utf-8") as f:
        data = json.load(f)
    resultado = {}
    for mes, unidades in data.get("MESES", {}).items():
        datos_mes = {}
        for unidad, vals in unidades.items():
            datos_mes[unidad] = (vals.get("numerador"), vals.get("denominador"), vals.get("%"))
        if datos_mes:
            resultado[mes] = datos_mes
    return resultado

hay_diferencias = False

for cat_dir in sorted(RUTA_HISTORIAL.iterdir()):
    if not cat_dir.is_dir() or cat_dir.name in EXCLUIR:
        continue
    for ind_dir in sorted(cat_dir.iterdir()):
        if not ind_dir.is_dir():
            continue
        indicador = ind_dir.name
        for ano_dir in sorted(ind_dir.iterdir()):
            if not ano_dir.is_dir() or not ano_dir.name.isdigit():
                continue
            ano = ano_dir.name
            excels = [f for f in ano_dir.glob("*.xlsx") if ano in f.stem]
            if not excels:
                continue

            ruta_excel = excels[0]
            nombre_json = indicador.replace(" ", "_") + ".json"
            ruta_json   = RUTA_DATA_FTP / ano / indicador.split()[0] / nombre_json

            datos_excel = leer_excel(ruta_excel)
            datos_json  = leer_json(ruta_json)

            meses_excel = set(datos_excel.keys())
            meses_json  = set(datos_json.keys())

            diferencias = []

            # Meses en Excel que no estan en JSON
            for mes in sorted(meses_excel - meses_json):
                diferencias.append(f"  MES FALTANTE en JSON: {mes} ({len(datos_excel[mes])} unidades)")

            # Unidades faltantes o datos distintos
            for mes in sorted(meses_excel & meses_json):
                unds_excel = set(datos_excel[mes].keys())
                unds_json  = set(datos_json[mes].keys())
                faltantes  = unds_excel - unds_json
                if faltantes:
                    diferencias.append(f"  {mes}: unidades faltantes en JSON: {sorted(faltantes)}")
                # Datos distintos
                for u in unds_excel & unds_json:
                    xe = datos_excel[mes][u]
                    xj = datos_json[mes][u]
                    if xe != xj:
                        diferencias.append(f"  {mes} / {u}: Excel={xe}  JSON={xj}")

            if diferencias:
                hay_diferencias = True
                print(f"\n[{indicador}] {ano}")
                for d in diferencias:
                    print(d)

if not hay_diferencias:
    print("Sin diferencias: todos los datos del Excel estan en el JSON.")

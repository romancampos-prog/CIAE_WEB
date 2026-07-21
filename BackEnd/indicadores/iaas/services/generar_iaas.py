"""
Genera el Excel completo IAAS (01-06) con históricos, semáforo y acumulados.
El cálculo (semáforo, acumulados, valores por celda) vive en calculos_iaas.py —
este archivo solo decide en qué fila/columna va cada valor y con qué estilo.
Usado en: iass/services/procesar_service.py, iass/controllers/reportes_controller.py
"""
import io
import xlsxwriter
import openpyxl
from iaas.config import MESES, RUTA_DATA_IAAS
from iaas.services.calculos_iaas import (
    _CFG, UNIDADES_IAAS, UNIDADES_UCI,
    _filas_umbrales_iaas01, _rango_umbral_uci,
    _leer_historicos_IAAS,
    calcular_fila_iaas01, calcular_acumulado_iaas01, calcular_anual_iaas01,
    calcular_fila_iaas_uci, calcular_acumulado_iaas_uci, calcular_anual_iaas_uci,
)

RUTA_BASE_IAAS = RUTA_DATA_IAAS


def _layout_iaas01():
    """
    Calcula, en un solo lugar, todas las filas clave de la hoja IAAS 01. Se deriva del número
    real de grupos de umbral (ver _filas_umbrales_iaas01), no de constantes fijas — así que
    _escribir_IAAS01, _escribir_datos_IAAS01 y _escribir_acumulado_IAAS01 siempre coinciden
    sin importar cuántas filas ocupe la tabla de umbrales.
    """
    N              = len(UNIDADES_IAAS)
    n_filas_umbral = len(_filas_umbrales_iaas01())

    fila_nota = 3

    # "MENSUAL" lleva el umbral (una fila por grupo) anclado bajo enero, metido entre el
    # renglón del nombre de mes y el encabezado — por eso +3+n_filas_umbral en vez de +3.
    fila_seccion_mensual = fila_nota + 1
    fila_inicio_mensual  = fila_seccion_mensual + 3 + n_filas_umbral
    fila_fin_mensual     = fila_inicio_mensual - 1 + N

    fila_seccion_acumulado = fila_fin_mensual + 2
    fila_inicio_acumulado  = fila_seccion_acumulado + 3
    fila_fin_acumulado     = fila_inicio_acumulado - 1 + N

    # Anual no tiene renglón de nombre de mes, por eso son 2 filas de encabezado y no 3.
    fila_seccion_anual = fila_fin_acumulado + 2
    fila_inicio_anual  = fila_seccion_anual + 2

    return {
        "fila_seccion_mensual":   fila_seccion_mensual,
        "fila_inicio_mensual":    fila_inicio_mensual,
        "fila_seccion_acumulado": fila_seccion_acumulado,
        "fila_inicio_acumulado":  fila_inicio_acumulado,
        "fila_seccion_anual":     fila_seccion_anual,
        "fila_inicio_anual":      fila_inicio_anual,
    }


def _escribir_unidades(hoja, estilos, fila_inicio, columna):
    for unidad in UNIDADES_IAAS:
        es_total = unidad == "DELEGACION"
        texto    = "OOAD" if es_total else unidad
        estilo   = estilos["delegacion_txt"] if es_total else estilos["lista_unidades_txt"]
        hoja.write(fila_inicio, columna, texto, estilo)
        fila_inicio += 1


def _escribir_bloque_mesIAAS_01(hoja, estilos, fila_titulo_mes, fila_subencabezado, col_inicio, cantidad_unidades, nombre_mes, con_fila_mes=True):
    if con_fila_mes:
        hoja.merge_range(fila_titulo_mes, col_inicio, fila_titulo_mes, col_inicio + 3, nombre_mes, estilos["encabezado_meses"])
    hoja.set_row(fila_subencabezado, 28)
    hoja.set_column(col_inicio,     col_inicio,     24)
    hoja.set_column(col_inicio + 1, col_inicio + 1, 12)
    hoja.set_column(col_inicio + 2, col_inicio + 2, 11)
    hoja.set_column(col_inicio + 3, col_inicio + 3, 11)
    hoja.write(fila_subencabezado, col_inicio,     "EGRESOS",            estilos["lista_unidades_txt"])
    hoja.write(fila_subencabezado, col_inicio + 1, "INFECCIONES",        estilos["lista_unidades_txt"])
    hoja.write(fila_subencabezado, col_inicio + 2, "DÍAS ESTANCIA",      estilos["lista_unidades_txt"])
    hoja.write(fila_subencabezado, col_inicio + 3, "TASA POR 1000 DÍAS", estilos["lista_unidades_txt"])

    for i in range(cantidad_unidades):
        estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
        for offset in range(4):
            hoja.write(fila_subencabezado + 1 + i, col_inicio + offset, "", estilo)

    fila_delegacion = fila_subencabezado + cantidad_unidades
    for offset in range(4):
        hoja.write(fila_delegacion, col_inicio + offset, "", estilos["delegacion_txt"])


def _escribir_IAAS01(libro, estilos):
    COLS_MES           = 4
    TOTAL_COLS         = 1 + 12 * COLS_MES
    cantidad_unidades  = len(UNIDADES_IAAS)
    layout             = _layout_iaas01()

    fila_titulo1 = 0
    fila_titulo2 = 1
    fila_titulo3 = 2
    fila_nota    = 3

    hoja = libro.add_worksheet("IAAS 01")
    hoja.set_column(0, 0, 27)
    hoja.set_default_row(20)
    hoja.freeze_panes(0, 1)
    libro.formats[0].set_font_name("Calibri")

    hoja.merge_range(fila_titulo1, 1, fila_titulo1, TOTAL_COLS - 1, "INFORMES DE INFECCIONES NOSOCOMIALES", estilos["titulo"])
    hoja.merge_range(fila_titulo2, 1, fila_titulo2, TOTAL_COLS - 1, "DELEGACIÓN GUANAJUATO",                estilos["titulo"])
    hoja.merge_range(fila_titulo3, 1, fila_titulo3, TOTAL_COLS - 1, "INFORMES DE INFECCIONES NOSOCOMIALES", estilos["titulo"])
    hoja.merge_range(fila_nota,    1, fila_nota,    TOTAL_COLS - 1,
        "NOTA:LA SUMATORIA INCLUYE UNICAMENTE LAS 8 UNIDADES QUE ENTRAN A EVALUACIÓN QUE SON LA 2, 3, 4, 10, 13, 21, 54, 58",
        estilos["titulo_tabla"])

    def _bloque(fila_seccion, etiqueta, meses, con_fila_mes=True, filas_umbral=None):
        """Escribe la etiqueta de sección (igual que IAAS 02-06) + UNIDADES + los meses dados.
        El ancho del bloque se ajusta a cuántos meses traiga — así "MENSUAL ACUMULADO" con
        11 meses (sin enero) no deja una columna de más ni una vacía. con_fila_mes=False quita
        el renglón del nombre de mes (bloque ANUAL, que no tiene mes que mostrar).
        filas_umbral (solo "MENSUAL"): inserta una fila por grupo de umbral (ver
        _filas_umbrales_iaas01) entre el nombre de mes y el encabezado, repetida debajo de
        cada mes del bloque."""
        if con_fila_mes:
            fila_encab = fila_seccion + 1
            siguiente  = fila_encab + 1
        else:
            fila_encab = fila_seccion
            siguiente  = fila_seccion + 1

        if filas_umbral:
            fila_umbral_inicio = siguiente
            fila_sub = fila_umbral_inicio + len(filas_umbral)
        else:
            fila_sub = siguiente

        ancho_bloque = 1 + len(meses) * COLS_MES
        hoja.write(fila_seccion, 0, "", estilos["titulo_iass06"])
        hoja.merge_range(fila_seccion, 1, fila_seccion, ancho_bloque - 1, etiqueta, estilos["titulo_iass06"])
        if con_fila_mes:
            hoja.merge_range(fila_encab, 0, fila_sub, 0, "UNIDADES", estilos["unidades_txt"])
        else:
            hoja.write(fila_sub, 0, "UNIDADES", estilos["unidades_txt"])

        _escribir_unidades(hoja, estilos, fila_sub + 1, 0)
        col = 1
        for mes in meses:
            _escribir_bloque_mesIAAS_01(hoja, estilos, fila_encab, fila_sub, col, cantidad_unidades, mes, con_fila_mes=con_fila_mes)
            if filas_umbral:
                for i, f in enumerate(filas_umbral):
                    fila = fila_umbral_inicio + i
                    hoja.write(fila, col,     f["etiqueta"], estilos["lista_unidades_txt"])
                    hoja.write(fila, col + 1, f["esperado"], estilos["Esperado_Leyenda"])
                    hoja.write(fila, col + 2, f["medio"],    estilos["Medio_Leyenda"])
                    hoja.write(fila, col + 3, f["bajo"],     estilos["Bajo_Leyenda"])
            col += COLS_MES

    # Bloque "MENSUAL": los 12 meses normales, uno arriba del otro con el resto de indicadores.
    # Lleva el umbral (una fila por grupo, ver _filas_umbrales_iaas01) repetido bajo cada mes.
    _bloque(layout["fila_seccion_mensual"], "MENSUAL", MESES, filas_umbral=_filas_umbrales_iaas01())

    # Bloque "MENSUAL ACUMULADO": debajo del mensual, sin enero — arranca directo en febrero,
    # no queda como una columna vacía al inicio.
    _bloque(layout["fila_seccion_acumulado"], "MENSUAL ACUMULADO", MESES[1:])

    # Bloque "ANUAL": al final, sin renglón de mes — directo del título a los datos.
    _bloque(layout["fila_seccion_anual"], "ANUAL", [""], con_fila_mes=False)


def Excel_IAAS01(anio: str = None):
    salida  = io.BytesIO()
    libro   = xlsxwriter.Workbook(salida)
    libro.set_properties({'author': 'Web CIAE'})
    estilos = Estilos_IAAS01(libro)
    _escribir_IAAS01(libro, estilos)

    if anio:
        hoja = libro.get_worksheet_by_name("IAAS 01")
        all_months = dict(_leer_historicos_IAAS(anio, 0).get("IAAS 01", {}))
        for m, mes_datos in all_months.items():
            _escribir_datos_IAAS01(hoja, estilos, m, mes_datos)
        _escribir_acumulado_IAAS01(hoja, estilos, all_months)
        _escribir_anual_IAAS01(hoja, estilos, all_months)

    libro.close()
    salida.seek(0)
    return salida


_COLOR_VERDE  = "#0B5445"
_COLOR_DORADO = "#9A7026"
_COLOR_ROJO   = "#7E0808"


def Estilos_IAAS01(libro):
    _cap = {"font_size": 9, "align": "center", "valign": "vcenter",
            "bold": True, "border": 1, "border_color": "#A6A6A6", "num_format": "0.00"}
    return {
        "titulo":               libro.add_format({"font_size": 9, "align": "center"}),
        "titulo_tabla":         libro.add_format({"font_size": 16, "align": "center", "bold": True}),
        "unidades_txt":         libro.add_format({"font_size": 11, "align": "center", "bold": True, "valign": "vcenter",
                                                   "bg_color": "#F2F2F2", "border_color": "#A6A6A6", "border": 1}),
        "lista_unidades_txt":   libro.add_format({"font_size": 9, "bg_color": "#F2F2F2", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "center", "text_wrap": True}),
        "delegacion_txt":       libro.add_format({"font_size": 9, "bg_color": "#808080", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "center", "text_wrap": True,
                                                   "bold": True, "font_color": "#FFFFFF"}),
        "encabezado_meses":     libro.add_format({"font_size": 11, "bg_color": "#F2F2F2", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "center", "bold": True,
                                                   "font_color": "#3D3D3D"}),
        "color_celda1":         libro.add_format({"font_size": 9, "bg_color": "#f9f9f9", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "center",
                                                   "text_wrap": True, "bold": True, "font_color": "#000000"}),
        "color_celda2":         libro.add_format({"font_size": 9, "bg_color": "#FFFFFF", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "center",
                                                   "text_wrap": True, "bold": True, "font_color": "#000000"}),
        "titulo_iass06":        libro.add_format({"font_size": 11, "bold": True, "align": "center", "valign": "vcenter",
                                                   "border": 1, "border_color": "#A6A6A6",
                                                   "bg_color": "#F2F2F2", "font_color": "#3D3D3D"}),
        "header_grupo2":        libro.add_format({"font_size": 9, "bg_color": "#808080", "border_color": "#A6A6A6",
                                                   "border": 1, "valign": "vcenter", "align": "left", "bold": True,
                                                   "font_color": "#FFFFFF", "text_wrap": True}),
        "Esperado": libro.add_format({**_cap, "bg_color": _COLOR_VERDE,  "font_color": "white"}),
        "Medio":    libro.add_format({**_cap, "bg_color": _COLOR_DORADO, "font_color": "white"}),
        "Bajo":     libro.add_format({**_cap, "bg_color": _COLOR_ROJO,   "font_color": "white"}),
        "sin_datos":libro.add_format({**_cap, "bg_color": "#CCCCCC",     "font_color": "black"}),
        # Igual que FTP: fondo gris, texto del color — no fondo sólido de color (esos son
        # para celdas de dato, estos son para el renglón de leyenda del umbral).
        "Esperado_Leyenda": libro.add_format({**_cap, "bg_color": "#F2F2F2", "font_color": _COLOR_VERDE}),
        "Medio_Leyenda":    libro.add_format({**_cap, "bg_color": "#F2F2F2", "font_color": _COLOR_DORADO}),
        "Bajo_Leyenda":     libro.add_format({**_cap, "bg_color": "#F2F2F2", "font_color": _COLOR_ROJO}),
    }


_IAAS_UCI_CONFIG = {
    num: {
        "indicador": f"IAAS {num}",
        "codigo":   _CFG[f"IAAS {num}"]["Titulo2"],
        "titulo":   _CFG[f"IAAS {num}"]["Titulo1"],
        "hoja":     f"IAAS {num}",
        "sub_col1": _CFG[f"IAAS {num}"]["subT1"],
        "sub_col2": _CFG[f"IAAS {num}"]["subT2"],
        "alto_sub": 48 if num == "04" else 30,
    }
    for num in ("02", "03", "04", "05", "06")
}


def _layout_iaas_uci():
    """
    Filas clave compartidas por _escribir_IAAS_UCI, _escribir_datos_IAAS_uci y
    _escribir_acumulado_IAAS_uci — un solo lugar, igual que _layout_iaas01, para que las
    tres funciones siempre coincidan (incluye la fila de OOAD de cada bloque).
    """
    N = len(UNIDADES_UCI)

    # "MENSUAL" lleva 3 filas de más (Esperado/Medio/Bajo) pegadas arriba de sus encabezados,
    # igual que FTP — por eso +6 en vez de +3. Acumulado y Anual no llevan umbral, se quedan igual.
    fila_inicio_mensual = 1
    fila_hosp1_mensual  = fila_inicio_mensual + 6
    fila_fin_mensual     = fila_inicio_mensual + N + 5
    fila_delega_mensual  = fila_fin_mensual + 1

    fila_inicio_acumulado = fila_delega_mensual + 2
    fila_hosp1_acumulado  = fila_inicio_acumulado + 3
    fila_fin_acumulado     = fila_inicio_acumulado + N + 2
    fila_delega_acumulado  = fila_fin_acumulado + 1

    # Anual no tiene renglón de nombre de mes (con_fila_mes=False), por eso son 2 filas de
    # encabezado en vez de 3 antes de los datos.
    fila_inicio_anual = fila_delega_acumulado + 2
    fila_hosp1_anual  = fila_inicio_anual + 2
    fila_fin_anual     = fila_inicio_anual + N + 1
    fila_delega_anual  = fila_fin_anual + 1

    return {
        "fila_inicio_mensual":   fila_inicio_mensual,
        "fila_hosp1_mensual":    fila_hosp1_mensual,
        "fila_delega_mensual":   fila_delega_mensual,
        "fila_inicio_acumulado": fila_inicio_acumulado,
        "fila_hosp1_acumulado":  fila_hosp1_acumulado,
        "fila_delega_acumulado": fila_delega_acumulado,
        "fila_inicio_anual":     fila_inicio_anual,
        "fila_hosp1_anual":      fila_hosp1_anual,
        "fila_delega_anual":     fila_delega_anual,
    }


def _escribir_tabla_iass_uci(hoja, estilos, fila_inicio, etiqueta_seccion, cfg, meses=None, con_fila_mes=True, con_umbral=False):
    """meses=None → los 12 completos (bloque MENSUAL). Pasar MESES[1:] para omitir
    enero por completo (bloque MENSUAL ACUMULADO) — el bloque se angosta con él.
    con_fila_mes=False quita el renglón del nombre de mes (bloque ANUAL, que no tiene
    mes que mostrar) — sin eso, sobraba un renglón vacío entre el título y los datos.
    con_umbral=True agrega 3 filas (Esperado/Medio/Bajo) pegadas arriba de los encabezados
    de cada mes, igual que hace FTP — solo tiene sentido en el bloque "MENSUAL"."""
    meses         = MESES if meses is None else meses
    COLS_MES      = 3
    ancho_bloque  = 1 + len(meses) * COLS_MES
    cant_unidades = len(UNIDADES_UCI)

    fila_seccion = fila_inicio
    if con_fila_mes:
        fila_mes  = fila_seccion + 1
        siguiente = fila_mes + 1
    else:
        fila_mes  = fila_seccion
        siguiente = fila_seccion + 1

    if con_umbral:
        fila_esperado = siguiente
        fila_medio    = fila_esperado + 1
        fila_bajo     = fila_medio + 1
        fila_sub      = fila_bajo + 1
    else:
        fila_sub = siguiente
    fila_hosp1 = fila_sub + 1

    hoja.write(fila_seccion, 0, "", estilos["titulo_iass06"])
    hoja.merge_range(fila_seccion, 1, fila_seccion, ancho_bloque - 1, etiqueta_seccion, estilos["titulo_iass06"])
    if con_fila_mes:
        hoja.merge_range(fila_mes, 0, fila_sub, 0, "UNIDADES", estilos["unidades_txt"])
    else:
        hoja.write(fila_sub, 0, "UNIDADES", estilos["unidades_txt"])
    hoja.set_row(fila_sub, cfg["alto_sub"])

    rango = _rango_umbral_uci(cfg["indicador"]) if con_umbral else None

    col = 1
    for mes in meses:
        if con_fila_mes:
            hoja.merge_range(fila_mes, col, fila_mes, col + 2, mes, estilos["encabezado_meses"])
        if con_umbral:
            hoja.merge_range(fila_esperado, col, fila_esperado, col + 2, f"ESPERADO: {rango['esperado']}", estilos["Esperado_Leyenda"])
            hoja.merge_range(fila_medio,    col, fila_medio,    col + 2, f"MEDIO: {rango['medio']}",       estilos["Medio_Leyenda"])
            hoja.merge_range(fila_bajo,     col, fila_bajo,     col + 2, f"BAJO: {rango['bajo']}",         estilos["Bajo_Leyenda"])
        hoja.write(fila_sub, col,     cfg["sub_col1"], estilos["lista_unidades_txt"])
        hoja.write(fila_sub, col + 1, cfg["sub_col2"], estilos["lista_unidades_txt"])
        hoja.write(fila_sub, col + 2, "Tasa",          estilos["lista_unidades_txt"])
        col += COLS_MES

    for i, unidad in enumerate(UNIDADES_UCI):
        fila   = fila_hosp1 + i
        estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
        hoja.write(fila, 0, unidad, estilos["lista_unidades_txt"])
        for c in range(1, ancho_bloque):
            hoja.write(fila, c, "", estilo)

    return fila_hosp1 + cant_unidades - 1


def _escribir_IAAS_UCI(libro, estilos, numero: str):
    cfg        = _IAAS_UCI_CONFIG[numero]
    COLS_MES   = 3
    TOTAL_COLS = 1 + 12 * COLS_MES
    layout     = _layout_iaas_uci()

    hoja = libro.add_worksheet(cfg["hoja"])
    hoja.set_column(0, 0, 32)
    hoja.freeze_panes(0, 1)
    hoja.set_column(1, TOTAL_COLS - 1, 9)
    for m in range(12):
        hoja.set_column(2 + m * COLS_MES, 2 + m * COLS_MES, 14)
    hoja.set_default_row(20)
    libro.formats[0].set_font_name("Calibri")

    hoja.merge_range(0, 0, 0, TOTAL_COLS - 1, cfg["titulo"], estilos["titulo_iass06"])

    def _fila_ooad(fila, ancho):
        hoja.write(fila, 0, "OOAD", estilos["delegacion_txt"])
        for c in range(1, ancho):
            hoja.write(fila, c, "", estilos["delegacion_txt"])

    _escribir_tabla_iass_uci(hoja, estilos, layout["fila_inicio_mensual"], "MENSUAL", cfg, meses=MESES, con_umbral=True)
    _fila_ooad(layout["fila_delega_mensual"], TOTAL_COLS)

    # "MENSUAL ACUMULADO" sin enero — el bloque arranca directo en febrero, más angosto.
    _escribir_tabla_iass_uci(hoja, estilos, layout["fila_inicio_acumulado"], "MENSUAL ACUMULADO", cfg, meses=MESES[1:])
    _fila_ooad(layout["fila_delega_acumulado"], 1 + len(MESES[1:]) * COLS_MES)

    # "ANUAL": sin renglón de mes (no tiene mes que mostrar) — directo del título a los datos.
    _escribir_tabla_iass_uci(hoja, estilos, layout["fila_inicio_anual"], "ANUAL", cfg, meses=[""], con_fila_mes=False)
    _fila_ooad(layout["fila_delega_anual"], 1 + COLS_MES)


def _Excel_IAAS_UCI(numero: str, anio: str = None) -> io.BytesIO:
    salida  = io.BytesIO()
    libro   = xlsxwriter.Workbook(salida)
    libro.set_properties({'author': 'Web CIAE'})
    estilos = Estilos_IAAS01(libro)
    _escribir_IAAS_UCI(libro, estilos, numero)

    if anio:
        clave      = f"IAAS {numero}"
        hoja       = libro.get_worksheet_by_name(clave)
        all_months = dict(_leer_historicos_IAAS(anio, 0).get(clave, {}))
        for m, mes_datos in all_months.items():
            _escribir_datos_IAAS_uci(hoja, estilos, m, mes_datos, clave)
        _escribir_acumulado_IAAS_uci(hoja, estilos, all_months, clave)
        _escribir_anual_IAAS_uci(hoja, estilos, all_months, clave)

    libro.close()
    salida.seek(0)
    return salida


def Excel_IAAS02(anio: str = None): return _Excel_IAAS_UCI("02", anio)
def Excel_IAAS03(anio: str = None): return _Excel_IAAS_UCI("03", anio)
def Excel_IAAS04(anio: str = None): return _Excel_IAAS_UCI("04", anio)
def Excel_IAAS05(anio: str = None): return _Excel_IAAS_UCI("05", anio)
def Excel_IAAS06(anio: str = None): return _Excel_IAAS_UCI("06", anio)


def _estilo_tasa(estilos, color):
    return estilos.get(color, estilos["sin_datos"])


def _escribir_acumulado_IAAS01(hoja, estilos, all_months):
    N = len(UNIDADES_IAAS)
    # Bloque "MENSUAL ACUMULADO": la fila se calcula del layout compartido, no es un número fijo.
    # Enero no tiene columna en este bloque (acumulado de enero = enero, no aporta nada nuevo) —
    # febrero es la primera columna, por eso el -2 en vez de -1.
    fila_inicio = _layout_iaas01()["fila_inicio_acumulado"]
    calculado   = calcular_acumulado_iaas01(all_months)

    for mes_target, fila_mes in calculado.items():
        col_base = 1 + (mes_target - 2) * 4

        for i, unidad in enumerate(UNIDADES_IAAS):
            if unidad == "DELEGACION":
                continue
            val = fila_mes[unidad]
            if val is None:
                continue
            fila   = fila_inicio + i
            estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
            hoja.write(fila, col_base + 1, val["numerador"],   estilo)
            hoja.write(fila, col_base + 2, val["denominador"], estilo)
            hoja.write(fila, col_base + 3, val["tasa"],        _estilo_tasa(estilos, val["color_tasa"]))

        fila_del = fila_inicio + N - 1
        val_del  = fila_mes["DELEGACION"]
        est_del  = estilos["delegacion_txt"]
        hoja.write(fila_del, col_base + 1, val_del["numerador"],   est_del)
        hoja.write(fila_del, col_base + 2, val_del["denominador"], est_del)
        hoja.write(fila_del, col_base + 3, val_del["tasa"],        _estilo_tasa(estilos, val_del["color_tasa"]))


def _escribir_anual_IAAS01(hoja, estilos, all_months):
    """Bloque 'ANUAL': el acumulado Ene→último mes registrado, la misma cuenta que ya
    hace _escribir_acumulado_IAAS01 para su última columna, pero puesta aparte para
    verse de un vistazo sin tener que ir hasta el final del bloque mensual acumulado."""
    calculado = calcular_anual_iaas01(all_months)
    if calculado is None:
        return
    N           = len(UNIDADES_IAAS)
    fila_inicio = _layout_iaas01()["fila_inicio_anual"]
    col_base    = 1

    for i, unidad in enumerate(UNIDADES_IAAS):
        if unidad == "DELEGACION":
            continue
        val = calculado[unidad]
        if val is None:
            continue
        fila   = fila_inicio + i
        estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
        hoja.write(fila, col_base + 1, val["numerador"],   estilo)
        hoja.write(fila, col_base + 2, val["denominador"], estilo)
        hoja.write(fila, col_base + 3, val["tasa"],        _estilo_tasa(estilos, val["color_tasa"]))

    fila_del = fila_inicio + N - 1
    val_del  = calculado["DELEGACION"]
    est_del  = estilos["delegacion_txt"]
    hoja.write(fila_del, col_base + 1, val_del["numerador"],   est_del)
    hoja.write(fila_del, col_base + 2, val_del["denominador"], est_del)
    hoja.write(fila_del, col_base + 3, val_del["tasa"],        _estilo_tasa(estilos, val_del["color_tasa"]))


def _escribir_acumulado_IAAS_uci(hoja, estilos, all_months, indicador):
    layout    = _layout_iaas_uci()
    fila_ini  = layout["fila_hosp1_acumulado"]
    fila_del  = layout["fila_delega_acumulado"]
    calculado = calcular_acumulado_iaas_uci(all_months, indicador)

    for mes_target, fila_mes in calculado.items():
        col_base = 1 + (mes_target - 2) * 3

        for i, unidad in enumerate(UNIDADES_UCI):
            val = fila_mes[unidad]
            if val is None:
                continue
            fila   = fila_ini + i
            estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
            hoja.write(fila, col_base,     val["numerador"],   estilo)
            hoja.write(fila, col_base + 1, val["denominador"], estilo)
            hoja.write(fila, col_base + 2, val["tasa"],        _estilo_tasa(estilos, val["color_tasa"]))

        val_ooad = fila_mes["OOAD"]
        est_del  = estilos["delegacion_txt"]
        hoja.write(fila_del, col_base,     val_ooad["numerador"],   est_del)
        hoja.write(fila_del, col_base + 1, val_ooad["denominador"], est_del)
        hoja.write(fila_del, col_base + 2, val_ooad["tasa"],        _estilo_tasa(estilos, val_ooad["color_tasa"]))


def _escribir_anual_IAAS_uci(hoja, estilos, all_months, indicador):
    """Bloque 'ANUAL': acumulado Ene→último mes registrado, mismo criterio que
    _escribir_anual_IAAS01 pero para IAAS 02-06 (OOAD es fila aparte, no de la lista)."""
    calculado = calcular_anual_iaas_uci(all_months, indicador)
    if calculado is None:
        return
    layout   = _layout_iaas_uci()
    fila_ini = layout["fila_hosp1_anual"]
    fila_del = layout["fila_delega_anual"]
    col_base = 1

    for i, unidad in enumerate(UNIDADES_UCI):
        val = calculado[unidad]
        if val is None:
            continue
        fila   = fila_ini + i
        estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
        hoja.write(fila, col_base,     val["numerador"],   estilo)
        hoja.write(fila, col_base + 1, val["denominador"], estilo)
        hoja.write(fila, col_base + 2, val["tasa"],        _estilo_tasa(estilos, val["color_tasa"]))

    val_ooad = calculado["OOAD"]
    est_del  = estilos["delegacion_txt"]
    hoja.write(fila_del, col_base,     val_ooad["numerador"],   est_del)
    hoja.write(fila_del, col_base + 1, val_ooad["denominador"], est_del)
    hoja.write(fila_del, col_base + 2, val_ooad["tasa"],        _estilo_tasa(estilos, val_ooad["color_tasa"]))


def _escribir_datos_IAAS_uci(hoja, estilos, mes_num, datos, indicador):
    layout   = _layout_iaas_uci()
    col_base = 1 + (mes_num - 1) * 3
    valores  = calcular_fila_iaas_uci(datos, indicador)

    for i, unidad in enumerate(UNIDADES_UCI):
        val    = valores[unidad]
        fila   = layout["fila_hosp1_mensual"] + i
        estilo = estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
        hoja.write(fila, col_base,     val["numerador"],   estilo)
        hoja.write(fila, col_base + 1, val["denominador"], estilo)
        hoja.write(fila, col_base + 2, val["tasa"],        _estilo_tasa(estilos, val["color_tasa"]))

    # Fila OOAD del bloque MENSUAL — antes no se calculaba (solo existía en el acumulado).
    val_ooad  = valores["OOAD"]
    fila_ooad = layout["fila_delega_mensual"]
    est_ooad  = estilos["delegacion_txt"]
    hoja.write(fila_ooad, col_base,     val_ooad["numerador"],   est_ooad)
    hoja.write(fila_ooad, col_base + 1, val_ooad["denominador"], est_ooad)
    hoja.write(fila_ooad, col_base + 2, val_ooad["tasa"],        _estilo_tasa(estilos, val_ooad["color_tasa"]))


def _escribir_datos_IAAS01(hoja, estilos, mes_num, datos):
    # Bloque "MENSUAL": la fila se calcula del layout compartido, no es un número fijo.
    fila_inicio = _layout_iaas01()["fila_inicio_mensual"]
    col_base    = 1 + (mes_num - 1) * 4
    valores     = calcular_fila_iaas01(datos)

    for i, unidad in enumerate(UNIDADES_IAAS):
        fila = fila_inicio + i
        val  = valores.get(unidad)
        if val is None:
            continue
        estilo = estilos["delegacion_txt"] if unidad == "DELEGACION" else (
            estilos["color_celda1"] if i % 2 == 0 else estilos["color_celda2"]
        )
        hoja.write(fila, col_base + 1, val["numerador"],   estilo)
        hoja.write(fila, col_base + 2, val["denominador"], estilo)
        hoja.write(fila, col_base + 3, val["tasa"],        _estilo_tasa(estilos, val["color_tasa"]))


def Excel_IAAS_Completo(anio: str, mes: str, datos: dict) -> io.BytesIO:
    mes_num      = int(mes)
    ruta_archivo = RUTA_BASE_IAAS / str(anio) / f"IAAS_{anio}.xlsx"
    historicos   = _leer_historicos_IAAS(anio, mes_num)

    salida  = io.BytesIO()
    libro   = xlsxwriter.Workbook(salida)
    libro.set_properties({'author': 'Web CIAE'})
    estilos = Estilos_IAAS01(libro)

    _escribir_IAAS01(libro, estilos)
    for num in ("02", "03", "04", "05", "06"):
        _escribir_IAAS_UCI(libro, estilos, num)

    hoja01 = libro.get_worksheet_by_name("IAAS 01")
    if hoja01:
        all_months_01 = dict(historicos.get("IAAS 01", {}))
        for m, mes_datos in all_months_01.items():
            _escribir_datos_IAAS01(hoja01, estilos, m, mes_datos)
        if datos.get("IAAS 01"):
            all_months_01[mes_num] = datos["IAAS 01"]
            _escribir_datos_IAAS01(hoja01, estilos, mes_num, datos["IAAS 01"])
        _escribir_acumulado_IAAS01(hoja01, estilos, all_months_01)
        _escribir_anual_IAAS01(hoja01, estilos, all_months_01)

    for num in ("02", "03", "04", "05", "06"):
        key  = f"IAAS {num}"
        hoja = libro.get_worksheet_by_name(key)
        if not hoja:
            continue

        all_months = dict(historicos.get(key, {}))
        if datos.get(key):
            all_months[mes_num] = datos[key]

        for m, mes_datos in all_months.items():
            _escribir_datos_IAAS_uci(hoja, estilos, m, mes_datos, key)

        _escribir_acumulado_IAAS_uci(hoja, estilos, all_months, key)
        _escribir_anual_IAAS_uci(hoja, estilos, all_months, key)

    libro.close()
    salida.seek(0)

    ruta_archivo.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(ruta_archivo, "wb") as f:
            f.write(salida.read())
        salida.seek(0)
        print(f"[IAAS] Guardado: {ruta_archivo}")
    except PermissionError:
        salida.seek(0)
        raise PermissionError(f"El archivo IAAS_{anio}.xlsx está abierto en Excel. Ciérralo e intenta de nuevo.")

    return salida